from selenium import webdriver
from selenium.webdriver.common.by import By
import time
import requests
import os
from selenium.webdriver.support.wait import WebDriverWait
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
import openpyxl
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.support import expected_conditions as EC

os.environ['PATH'] +=r"C:\Users\User\.spyder-py3"
chrome="C:\Program Files\Google\Chrome\Application\chromedriver.exe"

# Set your LinkedIn credentials
username = 'hashirpc60@gmail.com'
password = 'Xav58678'

driver=webdriver.Chrome()
driver.maximize_window()
wb = openpyxl.load_workbook('C:\\Users\\User\\Desktop\\uk linkedin.xlsx')
sh1 = wb['Sheet1']
# Open the LinkedIn login page
driver.get("https://www.linkedin.com/login")

# Find the username and password input fields, and enter your credentials
username_field = driver.find_element(By.ID, "username")
username_field.send_keys(username)

password_field = driver.find_element(By.ID, "password")
password_field.send_keys(password)

# Submit the login form
password_field.send_keys(Keys.RETURN)

for x in range(1184,1500):
    # Wait for the home page to load
    driver.implicitly_wait(5)
    data = sh1.cell(x+1,column=2)
    linkedinurl=data.value
    driver.get(data.value)
    
    #Name
    try:
         name=driver.find_element(By.CSS_SELECTOR,'.ember-view.text-display-medium-bold.org-top-card-summary__title.full-width') 
         print("Company Name: ",name.text)
         name=name.text
         sh1.cell(x+1, column=3, value=name)
    except:
        print("Name not found")
        name="Name not found"
        sh1.cell(x+1, column=3, value=name)
        
    #Services
    try:
        services=driver.find_element(By.CSS_SELECTOR,'div[class="block mt2"]')
        services=services.find_element(By.CSS_SELECTOR, '.org-top-card-summary-info-list__info-item') 
        print("Services: ",services.text)
        services=services.text
        sh1.cell(x+1, column=4, value=services)
    except:
        print("Services not found")
        services="Services not found"
        sh1.cell(x+1, column=4, value=services)
        
    #Location
    try:
        loc=driver.find_element(By.CSS_SELECTOR,'div[class="block mt2"]')
        loc=loc.find_element(By.CSS_SELECTOR,'div[class="inline-block"]')
        loc=loc.find_elements(By.CSS_SELECTOR, '.org-top-card-summary-info-list__info-item')
        print("Location: ",loc[0].text)
        locat=loc[0].text
        sh1.cell(x+1, column=5, value=locat)
    except:
        print("Location not found")   
        locat="Location not found"
        sh1.cell(x+1, column=5, value=locat)
        
    #followers
    try:
        follow=driver.find_element(By.CSS_SELECTOR,'div[class="block mt2"]')
        follow=follow.find_element(By.CSS_SELECTOR,'div[class="inline-block"]')
        follow=follow.find_elements(By.CSS_SELECTOR, '.org-top-card-summary-info-list__info-item')
        for item in follow:
            if "followers" in item.text:
                print("Followers: ",item.text)
                follow=item.text
                follow=follow.replace("followers","")
                sh1.cell(x+1, column=6, value=follow)
            
    except:
        print("Followers not found")
        follow="Followers not found"
        sh1.cell(x+1, column=6, value=follow)
        
    #Employees
    try:
        emp=driver.find_element(By.CSS_SELECTOR,'div[class="block mt2"]')
        emp=emp.find_element(By.CSS_SELECTOR,'div[class="inline-block"]')
        emp=emp.find_elements(By.CSS_SELECTOR, '.org-top-card-summary-info-list__info-item')
        for item in emp:
            if "employees" in item.text:
                print("Employees: ",item.text)
                emplo=item.text
                emplo=emplo.replace("employees", "")
                sh1.cell(x+1, column=7, value=emplo)
    except:
        print("Employees not found")
        emplo="Employees not found"
        sh1.cell(x+1, column=7, value=emplo)
        
        
    #Website URl
    try:
        website_element = driver.find_element(By.CSS_SELECTOR,'a[class="ember-view org-top-card-primary-actions__action"]')
        company_website_url = website_element.get_attribute('href')
        print("Company Website URL:", company_website_url)
        sh1.cell(x+1, column=8, value=company_website_url)

    except Exception as e:
        print("Unable to find the company website URL:", e)
        company_website_url="Company Url not Found"
        sh1.cell(x+1, column=8, value=company_website_url)
    
    
    # Get the HTML source code
    html = driver.page_source
    
    # Find the index of the specific word you're looking for
    word = "fsd_company:"
    index = html.find(word)
    
    if index != -1:
        # Extract the number following the specific word
        number_start_index = index + len(word)
        number_end_index = number_start_index
        while number_end_index < len(html) and html[number_end_index].isdigit():
            number_end_index += 1
        extracted_number = html[number_start_index:number_end_index]
        # Print the extracted number
        print("Organiztion Number:", extracted_number)
        sh1.cell(x+1, column=1, value=extracted_number)
        
        
    else:
        print("Organiztion Number not found")
        extracted_number="Organiztion Number not found"
        sh1.cell(x+1, column=1, value=extracted_number)
        
    
   

    #About info
    try:
        #Clicking About
        about=driver.find_element(By.CSS_SELECTOR,'.ember-view.full-width.text-align-center.link-without-hover-visited.org-module-card__footer-hoverable.pv4')
        about.click()
        driver.implicitly_wait(5)
        
        #Getting info
        about=driver.find_element(By.CSS_SELECTOR,'h2[class="text-heading-xlarge"]')
        about=WebDriverWait(driver, 10).until(EC.visibility_of_all_elements_located((By.CSS_SELECTOR, '.break-words.white-space-pre-wrap.t-black--light.text-body-medium')))  
        print("About: ",about[0].text)
        about_info=about[0].text
        sh1.cell(x+1, column=9, value=about_info)
        wb.save('C:\\Users\\User\\Desktop\\uk linkedin.xlsx')
        print(x," DONE")
    except:
        print("About Not found")
        about_info="About Not found"
        sh1.cell(x+1, column=9, value=about_info)
        wb.save('C:\\Users\\User\\Desktop\\uk linkedin.xlsx')
        print(x," DONE")
    
    # Prepare the data to be sent in the POST request
    payload = {
        'key':'anohtercrm_master_db',
        'company_org_id': extracted_number,
        'linkedin_url': linkedinurl,
        'company_name': name,
        'company_type': services,
        'company_location': locat,
        'company_follower': follow,
        'company_employee': emplo,
        'company_website': company_website_url,
        'company_info': about_info
    }
    
    # Send the POST request to the API endpoint
    api_url = 'https://app.anothercrm.com/api/store-mata-data'
    response = requests.post(api_url, json=payload)
    
    # Check the response status
    if response.status_code == 200:
        print('Data successfully posted to the API.')
    else:
        print('Failed to post data to the API.')

    
driver.close()

        

        