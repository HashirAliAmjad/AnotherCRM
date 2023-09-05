import re
import requests
from bs4 import BeautifulSoup
import openpyxl

# Define technology patterns and associated metadata
technology_patterns = {
    'WordPress': {
        'pattern': r'wp-content|wp-json|wordpress',
        'category': 'Content Management System',
    },
    'Joomla': {
        'pattern': r'joomla',
        'category': 'Content Management System',
    },
    'Drupal': {
        'pattern': r'drupal|sites\/all',
        'category': 'Content Management System',
    },
    'Magento': {
        'pattern': r'magento',
        'category': 'E-commerce',
    },
    'Shopify': {
        'pattern': r'shopify',
        'category': 'E-commerce',
    },
    'WooCommerce': {
        'pattern': r'woocommerce',
        'category': 'E-commerce',
    },
    'React': {
        'pattern': r'react',
        'category': 'JavaScript Framework',
    },
    'Angular': {
        'pattern': r'angular',
        'category': 'JavaScript Framework',
    },
    'Vue.js': {
        'pattern': r'vue\.js',
        'category': 'JavaScript Framework',
    },
    'Node.js': {
        'pattern': r'node\.js',
        'category': 'Server-Side JavaScript',
    },
    'Ruby on Rails': {
        'pattern': r'ruby on rails|rails',
        'category': 'Web Framework',
    },
    'Django': {
        'pattern': r'django',
        'category': 'Web Framework',
    },
    'Laravel': {
        'pattern': r'laravel',
        'category': 'Web Framework',
    },
    'Bootstrap': {
        'pattern': r'bootstrap',
        'category': 'Front-End Framework',
    },
    'jQuery': {
        'pattern': r'jquery',
        'category': 'JavaScript Library',
    },
    'TensorFlow': {
        'pattern': r'tensorflow',
        'category': 'Machine Learning',
    },
    'PyTorch': {
        'pattern': r'pytorch',
        'category': 'Machine Learning',
    },
    'Flask': {
        'pattern': r'flask',
        'category': 'Python Web Framework',
    },
    'Express.js': {
        'pattern': r'express\.js',
        'category': 'JavaScript Web Framework',
    },
    'ASP.NET': {
        'pattern': r'asp\.net',
        'category': 'Web Framework',
    },
    'Java Servlet': {
        'pattern': r'javax\.servlet',
        'category': 'Java Web Technology',
    },
    'Ruby': {
        'pattern': r'ruby',
        'category': 'Programming Language',
    },
    'Go': {
        'pattern': r'golang|go-lang',
        'category': 'Programming Language',
    },
    'MongoDB': {
        'pattern': r'mongodb',
        'category': 'Database',
    },
    'MySQL': {
        'pattern': r'mysql',
        'category': 'Database',
    },
    'PostgreSQL': {
        'pattern': r'postgresql|postgres',
        'category': 'Database',
    },
    'AWS': {
        'pattern': r'aws\.amazon|amazon web services',
        'category': 'Cloud Services',
    },
    'Google Cloud': {
        'pattern': r'google cloud|gcp',
        'category': 'Cloud Services',
    },
    'Heroku': {
        'pattern': r'heroku',
        'category': 'Platform as a Service',
    },
    'Firebase': {
        'pattern': r'firebase',
        'category': 'Backend as a Service',
    },
    'Elasticsearch': {
        'pattern': r'elasticsearch',
        'category': 'Search Engine',
    },
    'Redis': {
        'pattern': r'redis',
        'category': 'In-Memory Data Store',
    },
    'Apache': {
        'pattern': r'apache',
        'category': 'Web Server',
    },
    'Nginx': {
        'pattern': r'nginx',
        'category': 'Web Server',
    },
    'Docker': {
        'pattern': r'docker',
        'category': 'Containerization',
    },
    'Kubernetes': {
        'pattern': r'kubernetes',
        'category': 'Container Orchestration',
    },
    'Swift': {
        'pattern': r'swift',
        'category': 'iOS Development',
    },
    'Android': {
        'pattern': r'android',
        'category': 'Android Development',
    },
    'Unity': {
        'pattern': r'unity',
        'category': 'Game Development',
    },
    'Adobe Photoshop': {
        'pattern': r'adobe photoshop|photoshop',
        'category': 'Graphic Design',
    },
    'Adobe Illustrator': {
        'pattern': r'adobe illustrator|illustrator',
        'category': 'Graphic Design',
    },
    'Microsoft Office': {
        'pattern': r'microsoft office',
        'category': 'Productivity Suite',
    },
    'Google Workspace': {
        'pattern': r'google workspace',
        'category': 'Productivity Suite',
    },
    'Git': {
        'pattern': r'git',
        'category': 'Version Control System',
    },
    'Bitbucket': {
        'pattern': r'bitbucket',
        'category': 'Code Collaboration Platform',
    },
    'Jenkins': {
        'pattern': r'jenkins',
        'category': 'Continuous Integration and Continuous Deployment',
    },
    'Travis CI': {
        'pattern': r'travis ci|travis-ci',
        'category': 'Continuous Integration and Continuous Deployment',
    },
    'CircleCI': {
        'pattern': r'circle ci|circleci',
        'category': 'Continuous Integration and Continuous Deployment',
    },
    'Ansible': {
        'pattern': r'ansible',
        'category': 'Infrastructure Automation',
    },
    'Terraform': {
        'pattern': r'terraform',
        'category': 'Infrastructure as Code',
    },
    'Apache Kafka': {
        'pattern': r'apache kafka|kafka',
        'category': 'Distributed Streaming Platform',
    },
    'GraphQL': {
        'pattern': r'graphql',
        'category': 'Query Language for APIs',
    },
    'REST': {
        'pattern': r'rest|restful',
        'category': 'API Design',
    },
    'OAuth': {
        'pattern': r'oauth',
        'category': 'Authentication Protocol',
    },
    'JWT': {
        'pattern': r'jwt',
        'category': 'Authentication Token',
    },
    'OAuth2': {
        'pattern': r'oauth2',
        'category': 'Authentication Protocol',
    },
    'Swagger': {
        'pattern': r'swagger',
        'category': 'API Documentation',
    },
    'OpenAPI': {
        'pattern': r'openapi',
        'category': 'API Specification',
    },
    'Selenium': {
        'pattern': r'selenium',
        'category': 'Automated Browser Testing',
    },
    'Cucumber': {
        'pattern': r'cucumber',
        'category': 'Behavior-Driven Development',
    },
    'JUnit': {
        'pattern': r'junit',
        'category': 'Java Unit Testing Framework',
    },
    'Pytest': {
        'pattern': r'pytest',
        'category': 'Python Testing Framework',
    },
    'Jest': {
        'pattern': r'jest',
        'category': 'JavaScript Testing Framework',
    },
    'Mocha': {
        'pattern': r'mocha',
        'category': 'JavaScript Testing Framework',
    },
    'PHPUnit': {
        'pattern': r'phpunit',
        'category': 'PHP Testing Framework',
    },
    'C#': {
        'pattern': r'c#',
        'category': 'Programming Language',
    },
    'VB.NET': {
        'pattern': r'vb\.net',
        'category': 'Programming Language',
    },
    'Rust': {
        'pattern': r'rust',
        'category': 'Programming Language',
    },
    'Hadoop': {
        'pattern': r'hadoop',
        'category': 'Big Data Processing',
    },
    'Apache Spark': {
        'pattern': r'apache spark|spark',
        'category': 'Big Data Processing',
    },
    'Tableau': {
        'pattern': r'tableau',
        'category': 'Data Visualization',
    },
    'Power BI': {
        'pattern': r'power bi',
        'category': 'Data Visualization',
    },
    'Eclipse': {
        'pattern': r'eclipse',
        'category': 'Integrated Development Environment',
    },
    'Visual Studio Code': {
        'pattern': r'visual studio code|vs code',
        'category': 'Integrated Development Environment',
    },
    'Atom': {
        'pattern': r'atom',
        'category': 'Text Editor',
    },
    'Sublime Text': {
        'pattern': r'sublime text',
        'category': 'Text Editor',
    },
    'Adobe Premiere Pro': {
        'pattern': r'adobe premiere pro|premiere pro',
        'category': 'Video Editing',
    },
    'Final Cut Pro': {
        'pattern': r'final cut pro',
        'category': 'Video Editing',
    },
    'AutoCAD': {
        'pattern': r'autocad',
        'category': 'Computer-Aided Design',
    },
    'MATLAB': {
        'pattern': r'matlab',
        'category': 'Numerical Computing',
    },
    'Unreal Engine': {
        'pattern': r'unreal engine',
        'category': 'Game Development',
    },
    'Adobe After Effects': {
        'pattern': r'adobe after effects|after effects',
        'category': 'Motion Graphics',
    },
    'Blender': {
        'pattern': r'blender',
        'category': '3D Modeling and Animation',
    },
    'Arduino': {
        'pattern': r'arduino',
        'category': 'Hardware Programming',
    },
    'Raspberry Pi': {
        'pattern': r'raspberry pi',
        'category': 'Single-Board Computer',
    },
    'WordPress Plugin': {
    'pattern': r'wp-content/plugins',
    'category': 'WordPress Plugin',
    },
    'Drupal Module': {
        'pattern': r'drupal\/modules',
        'category': 'Drupal Module',
    },
    'Node.js Frameworks': {
        'pattern': r'express|koa|hapi|sails',
        'category': 'Node.js Framework',
    },
    'Ruby Gems': {
        'pattern': r'gemfile',
        'category': 'Ruby Gem',
    },
    'Python Packages': {
        'pattern': r'requirements\.txt',
        'category': 'Python Package',
    },
    'Java Frameworks': {
        'pattern': r'spring|struts|javaee',
        'category': 'Java Framework',
    },
    '.NET Framework': {
        'pattern': r'\.net|dotnet',
        'category': '.NET Framework',
    },
    'Vue.js UI Frameworks': {
        'pattern': r'vuetify|quasar|element ui',
        'category': 'Vue.js UI Framework',
    },
    'Front-End Libraries': {
        'pattern': r'react-dom|angular|vue',
        'category': 'Front-End Library',
    },
    'PHP Frameworks': {
        'pattern': r'laravel|codeigniter|symfony|cakephp',
        'category': 'PHP Framework',
    },
    'JavaScript Bundlers': {
        'pattern': r'webpack|rollup|parcel|gulp',
        'category': 'JavaScript Bundler',
    },
    'Mobile App Development': {
        'pattern': r'react native|flutter|ionic|cordova',
        'category': 'Mobile App Development',
    },
    'NoSQL Databases': {
        'pattern': r'mongodb|cassandra|couchbase',
        'category': 'NoSQL Database',
    },
    'SQL Server': {
        'pattern': r'sql server',
        'category': 'Database',
    },
    'GraphQL Libraries': {
        'pattern': r'apollo|graphql-js',
        'category': 'GraphQL Library',
    },
    'Containerization Tools': {
        'pattern': r'docker|kubernetes|rkt',
        'category': 'Containerization Tool',
    },
    'Game Engines': {
        'pattern': r'unity|unreal engine|godot',
        'category': 'Game Engine',
    },
    'Data Science Libraries': {
        'pattern': r'pandas|numpy|scikit-learn',
        'category': 'Data Science Library',
    },
    'CAD Software': {
        'pattern': r'autocad|solidworks|fusion 360',
        'category': 'CAD Software',
    },
    'API Gateways': {
        'pattern': r'kong|tyk|apigee',
        'category': 'API Gateway',
    },
    'UI/UX Design Tools': {
        'pattern': r'sketch|figma|invision',
        'category': 'UI/UX Design Tool',
    },
    'React Native': {
        'pattern': r'react native',
        'category': 'Mobile App Development',
    },
    'Flutter': {
        'pattern': r'flutter',
        'category': 'Mobile App Development',
    },
    'Ionic': {
        'pattern': r'ionic',
        'category': 'Mobile App Development',
    },
    'Cordova': {
        'pattern': r'cordova',
        'category': 'Mobile App Development',
    },
    'GraphQL-JS': {
        'pattern': r'graphql-js',
        'category': 'GraphQL Library',
    },
    'Apollo Client': {
        'pattern': r'apollo client',
        'category': 'GraphQL Library',
    },
    'Docker Compose': {
        'pattern': r'docker-compose',
        'category': 'Container Orchestration',
    },
    'RKT': {
        'pattern': r'rkt',
        'category': 'Containerization Tool',
    },
    'Godot': {
        'pattern': r'godot',
        'category': 'Game Engine',
    },
    'Pandas': {
        'pattern': r'pandas',
        'category': 'Data Science Library',
    },
    'NumPy': {
        'pattern': r'numpy',
        'category': 'Data Science Library',
    },
    'scikit-learn': {
        'pattern': r'scikit-learn',
        'category': 'Data Science Library',
    },
    'Fusion 360': {
        'pattern': r'fusion 360',
        'category': 'CAD Software',
    },
    'Kong': {
        'pattern': r'kong',
        'category': 'API Gateway',
    },
    'Tyk': {
        'pattern': r'tyk',
        'category': 'API Gateway',
    },
    'Apigee': {
        'pattern': r'apigee',
        'category': 'API Gateway',
    },
    'Sketch': {
        'pattern': r'sketch',
        'category': 'UI/UX Design Tool',
    },
    'Figma': {
        'pattern': r'figma',
        'category': 'UI/UX Design Tool',
    },
    'InVision': {
        'pattern': r'invision',
        'category': 'UI/UX Design Tool',
    },
    'Kotlin': {
        'pattern': r'kotlin',
        'category': 'Programming Language',
    },
    'RabbitMQ': {
        'pattern': r'rabbitmq',
        'category': 'Message Broker',
    },

    'Apache Solr': {
        'pattern': r'apache solr',
        'category': 'Search Engine',
    },

    'Memcached': {
        'pattern': r'memcached',
        'category': 'In-Memory Data Store',
    },
    'Apache Cassandra': {
        'pattern': r'apache cassandra',
        'category': 'NoSQL Database',
    },
    'Elixir': {
        'pattern': r'elixir',
        'category': 'Programming Language',
    },
    'VuePress': {
        'pattern': r'vuepress',
        'category': 'Static Site Generator',
    },
    'Kafka Streams': {
        'pattern': r'kafka streams',
        'category': 'Stream Processing',
    },
    'Elastic Stack': {
        'pattern': r'elastic stack',
        'category': 'Data Analytics',
    },
    'Salesforce': {
        'pattern': r'salesforce',
        'category': 'Customer Relationship Management',
    },
    'Salesforce Service Cloud': {
        'pattern': r'salesforce service cloud',
        'category': 'Customer Service',
    },
    'Salesforce Marketing Cloud Email Studio': {
        'pattern': r'salesforce marketing cloud email studio',
        'category': 'Marketing Automation',
    },
    'Salesforce Marketing Cloud Account Engagement': {
        'pattern': r'salesforce marketing cloud account engagement',
        'category': 'Marketing Automation',
    },
    'Salesforce Interaction Studio': {
        'pattern': r'salesforce interaction studio',
        'category': 'Customer Engagement',
    },
    'Salesforce Desk': {
        'pattern': r'salesforce desk',
        'category': 'Customer Support',
    },
    'Salesforce Commerce Cloud': {
        'pattern': r'salesforce commerce cloud',
        'category': 'E-commerce',
    },
    'Salesforce Audience Studio': {
        'pattern': r'salesforce audience studio',
        'category': 'Audience Management',
    },
}

def detect_technologies(url):
    try:
        response = requests.get(url)
        response.raise_for_status()  # Raise an exception for HTTP errors
        html_content = response.text

        detected_technologies = []

        # Parse HTML content
        soup = BeautifulSoup(html_content, 'html.parser')

        # Iterate through technology patterns
        for technology, tech_data in technology_patterns.items():
            pattern = tech_data['pattern']
            category = tech_data['category']

            if re.search(pattern, str(soup), re.IGNORECASE):
                detected_technologies.append({
                    'technology': technology,
                    'category': category
                })

        return detected_technologies
    except requests.exceptions.RequestException:
        print("Exception")
        return []

# Example usage
wb = openpyxl.load_workbook('C:\\Users\\User\\Desktop\\uni.xlsx')
sh1 = wb['Sheet1']

for x in range(1552, 1930):
    data = sh1.cell(x + 1, column=3)
    url = data.value
    detected_technologies = detect_technologies(url)
    tech_info = ""
    for tech in detected_technologies:
        print(f"Technology: {tech['technology']}, Category: {tech['category']}")
        tech_info += f"Technology: {tech['technology']}, Category: {tech['category']}\n"
    sh1.cell(x + 1, column=4, value=tech_info.strip())
    wb.save('C:\\Users\\User\\Desktop\\uni.xlsx')
    print(x,"Done")
