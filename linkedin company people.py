from selenium import webdriver
from selenium.webdriver.common.by import By
import time
import requests
import os
import json
import re
from bs4 import BeautifulSoup
from urllib.parse import urljoin
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
import openpyxl
from selenium.common.exceptions import NoSuchElementException
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.support import expected_conditions as EC

os.environ['PATH'] +=r"C:\Users\User\.spyder-py3"
chrome="C:\Program Files\Google\Chrome\Application\chromedriver.exe"

wb = openpyxl.load_workbook('C:\\Users\\User\\Desktop\\people1.xlsx')
sh1 = wb['Sheet1']

def login():
    # Set your LinkedIn credentials
    username = 'redrose12677@gmail.com'
    password = 'Xav58678'
    
    # Find the username and password input fields, and enter your credentials
    username_field = driver.find_element(By.ID, "username")
    username_field.send_keys(username)

    password_field = driver.find_element(By.ID, "password")
    password_field.send_keys(password)

    # Submit the login form
    password_field.send_keys(Keys.RETURN)
    
def extract_linkedin_details(linkedin_url, target_companies):
    # Open a new tab and switch to it
    driver.execute_script("window.open('');")
    driver.switch_to.window(driver.window_handles[1])

    # Navigate to the LinkedIn profile URL
    driver.get(linkedin_url)
    driver.implicitly_wait(7)
    #time.sleep(7)
    # Extract the name
    try:
        name="name"
        nm=driver.find_element(By.CSS_SELECTOR,'.scaffold-layout__row.scaffold-layout__content.scaffold-layout__content--main-aside.scaffold-layout__content--has-aside')
        nm=nm.find_element(By.CSS_SELECTOR,'main[class="scaffold-layout__main"]')
        nm=nm.find_element(By.CSS_SELECTOR,'section[class="artdeco-card ember-view pv-top-card"]')
        nm=nm.find_element(By.CSS_SELECTOR,'h1[class="text-heading-xlarge inline t-24 v-align-middle break-words"]')
        name = nm.text
        print("Name: ",name,"\n")
    except:
        name="None"
        print("Name: ",name,"\n")
        
    # Extract the contact information if available                   
    try:
        contact_info="none"
        cn=driver.find_element(By.CSS_SELECTOR,'.scaffold-layout__row.scaffold-layout__content.scaffold-layout__content--main-aside.scaffold-layout__content--has-aside')
        cn=cn.find_element(By.CSS_SELECTOR,'main[class="scaffold-layout__main"]')
        cn=cn.find_element(By.CSS_SELECTOR,'section[class="artdeco-card ember-view pv-top-card"]')
        cn=cn.find_element(By.CSS_SELECTOR,'a[class="ember-view link-without-visited-state cursor-pointer text-heading-small inline-block break-words"]')   
        cn.click()         
        time.sleep(1)
        cn=driver.find_element(By.CSS_SELECTOR,'div[class="artdeco-modal__content ember-view"]')
        contact_info=cn.text
        print("Contact Info: ",contact_info,"\n")
        cn=driver.find_element(By.CSS_SELECTOR,'button[class="artdeco-modal__dismiss artdeco-button artdeco-button--circle artdeco-button--muted artdeco-button--2 artdeco-button--tertiary ember-view"]')
        cn.click()
    except NoSuchElementException:
        contact_info="None"
        print("Contact Info: ",contact_info,"\n")
        
    # Extract the current job title and company
    try:
        current_job="none"
        current_company="None"
        jobs=driver.find_element(By.CSS_SELECTOR,'.scaffold-layout__row.scaffold-layout__content.scaffold-layout__content--main-aside.scaffold-layout__content--has-aside')
        jobs=jobs.find_element(By.CSS_SELECTOR,'main[class="scaffold-layout__main"]')
        jobs=jobs.find_elements(By.CSS_SELECTOR,'.artdeco-card.ember-view.relative.break-words.pb3')
        for job in jobs:
            jb=job.find_element(By.CSS_SELECTOR,'div[class="pvs-header__container"]')
            jb=jb.find_element(By.CSS_SELECTOR,'div[class="pvs-header__title-container"]')
            if "Experience" in jb.text:
                #Current Job
                driver.execute_script("arguments[0].scrollIntoView({ behavior: 'smooth', block: 'center' });", jb)
                jobe=job.find_element(By.CSS_SELECTOR,'div[class="pvs-list__outer-container"]')
                jobe=jobe.find_element(By.CSS_SELECTOR,'.pvs-list')
                jobe=jobe.find_element(By.CSS_SELECTOR,'li[class="artdeco-list__item pvs-list__item--line-separated pvs-list__item--one-column"]')
                jaaab=jobe.find_element(By.CSS_SELECTOR,'div[class="display-flex flex-column full-width align-self-center"]')
                jaaab=jaaab.find_element(By.CSS_SELECTOR,'div[class="display-flex flex-row justify-space-between"]')
                try:
                    jaaab.find_element(By.CSS_SELECTOR,'.optional-action-target-wrapper') 
                    job2=jobe.find_element(By.CSS_SELECTOR,'div[class="display-flex flex-column full-width align-self-center"]')
                    job2=job2.find_element(By.CSS_SELECTOR,'div[class="pvs-list__outer-container"]')
                    job2=job2.find_element(By.CSS_SELECTOR,'span[aria-hidden="true"]')
                    current_job=job2.text
                    print("Current Job: ",current_job,"\n")
                    
                    #Current Company
                    jobea=jobe.find_element(By.CSS_SELECTOR,'span[aria-hidden="true"]')
                    current_Company=jobea.text
                    print("Current Company: ",current_Company,"\n")
                       
                # Check the count of direct child div elements
                except NoSuchElementException:
                    jobea=jobe.find_element(By.CSS_SELECTOR,'span[aria-hidden="true"]')
                    current_job=jobea.text
                    print("Current Job: ",current_job,"\n")
                    
                    #Current Comany
                    jobeb=jobe.find_element(By.CSS_SELECTOR,'span[class="t-14 t-normal"]')
                    jobeb=jobeb.find_element(By.CSS_SELECTOR,'span[aria-hidden="true"]')
                    current_company=jobeb.text
                    print("Current Company: ",current_company,"\n")
    except NoSuchElementException:
         current_job="None"
         print("Current Job: ",current_job,"\n") 
         current_company="None"
         print("Current Company: ",current_company,"\n")              
                
            
    # Extract the about data
    try:
        about_data="None"
        jobs=driver.find_element(By.CSS_SELECTOR,'.scaffold-layout__row.scaffold-layout__content.scaffold-layout__content--main-aside.scaffold-layout__content--has-aside')
        jobs=jobs.find_element(By.CSS_SELECTOR,'main[class="scaffold-layout__main"]')
        jobs=jobs.find_elements(By.CSS_SELECTOR,'.artdeco-card.ember-view.relative.break-words.pb3.mt2')
        for job in jobs:
            jb=job.find_element(By.CSS_SELECTOR,'div[class="pvs-header__container"]')
            jb=jb.find_element(By.CSS_SELECTOR,'div[class="pvs-header__title-container"]')
            if "About" in jb.text:
                driver.execute_script("arguments[0].scrollIntoView({ behavior: 'smooth', block: 'center' });", jb)
                abt=job.find_element(By.CSS_SELECTOR,'div[class="display-flex ph5 pv3"]')
                abt=abt.find_element(By.CSS_SELECTOR,'span[aria-hidden="true"]')
                about_data=abt.text
                print("About Data: ",about_data,"\n")
    except:
        about_data="None"
        print("About Data: ",about_data,"\n")

    # Extract past job positions in specific companies
    past_job="None"
    past_date="None"
    past_company="None"
    try:
        print("0")
        exp=driver.find_element(By.CSS_SELECTOR,'a[id="navigation-index-see-all-experiences"]')
        driver.execute_script("arguments[0].scrollIntoView({ behavior: 'smooth', block: 'center' });", exp)
        WebDriverWait(driver, 10).until(EC.presence_of_all_elements_located)
        exp.click()
        jobs=driver.find_element(By.CSS_SELECTOR,'.scaffold-layout__row.scaffold-layout__content.scaffold-layout__content--main-aside.scaffold-layout__content--has-aside')
        jobs=jobs.find_element(By.CSS_SELECTOR,'main[class="scaffold-layout__main"]')
        #jobs=jobs.find_element(By.CSS_SELECTOR,'.artdeco-card.ember-view.relative.break-words.pb3')
        jobe=jobs.find_elements(By.CSS_SELECTOR,'li[class="pvs-list__paged-list-item artdeco-list__item pvs-list__item--line-separated pvs-list__item--one-column"]')
        for item in jobe:
            print("4")
            jaaab=item.find_element(By.CSS_SELECTOR,'div[class="display-flex flex-column full-width align-self-center"]')
            jaaab=jaaab.find_element(By.CSS_SELECTOR,'div[class="display-flex flex-row justify-space-between"]')
            print("5")
            try:
                jaaab=jaaab.find_element(By.CSS_SELECTOR,'.optional-action-target-wrapper')
                pjob=item.find_element(By.CSS_SELECTOR,'span[aria-hidden="true"]')
                print(pjob.text," :Line 177 Companies\n")
                if pjob.text in target_companies:
                    #past position
                   pjobea=item.find_element(By.CSS_SELECTOR,'div[class="pvs-list__outer-container"]')
                   pjobea=pjobea.find_element(By.CSS_SELECTOR,'span[aria-hidden="true"]')
                   past_job=pjobea.text 
                   print("Past Position",past_job,"\n")
                   dt=item.find_element(By.CSS_SELECTOR,'div[class="pvs-list__outer-container"]')
                   dt=dt.find_element(By.CSS_SELECTOR,'span[class="t-14 t-normal t-black--light"]')
                   past_date=dt.text
                   print("Past Date: ",past_date,"\n")
                   past_company=pjob.text
                   print("Past Company: ",past_company,"\n")
            except NoSuchElementException:
                job3=item.find_element(By.CSS_SELECTOR,'span[class="t-14 t-normal"]')
                job3=job3.find_element(By.CSS_SELECTOR,'span[aria-hidden="true"]')
                if job3.text in target_companies:
                   job4=item.find_element(By.CSS_SELECTOR,'span[aria-hidden="true"]')
                   past_job=job4.text
                   print("Past Job: ",past_job,"\n")
                   dt=item.find_element(By.CSS_SELECTOR,'span[class="t-14 t-normal t-black--light"]')
                   dt=dt.find_element(By.CSS_SELECTOR,'span[aria-hidden="true"]')
                   past_date=dt.text
                   print("Past Date: ",past_date,"\n")
                   past_company=job3.text
                   print("Past Company: ",past_company,"\n")
    except TimeoutException:
        print("Some Experience")
        try:
            print("1")
            jobs=driver.find_element(By.CSS_SELECTOR,'.scaffold-layout__row.scaffold-layout__content.scaffold-layout__content--main-aside.scaffold-layout__content--has-aside')
            jobs=jobs.find_element(By.CSS_SELECTOR,'main[class="scaffold-layout__main"]')
            jobs=jobs.find_elements(By.CSS_SELECTOR,'.artdeco-card.ember-view.relative.break-words.pb3')
            print("2")
            for job in jobs:
                print("3")
                jb=job.find_element(By.CSS_SELECTOR,'div[class="pvs-header__container"]')
                jb=jb.find_element(By.CSS_SELECTOR,'div[class="pvs-header__title-container"]')
                print("before if")
                if "Experience" in jb.text:
                    print("after if")
                    driver.execute_script("arguments[0].scrollIntoView({ behavior: 'smooth', block: 'center' });", jb)
                    jobe=job.find_element(By.CSS_SELECTOR,'div[class="pvs-list__outer-container"]')
                    jobe=jobe.find_elements(By.CSS_SELECTOR,'li[class="artdeco-list__item pvs-list__item--line-separated pvs-list__item--one-column pvs-list--ignore-first-item-top-padding"]')
                    for item in jobe:
                        print("4")
                        jaaab=item.find_element(By.CSS_SELECTOR,'div[class="display-flex flex-column full-width align-self-center"]')
                        jaaab=jaaab.find_element(By.CSS_SELECTOR,'div[class="display-flex flex-row justify-space-between"]')
                        print("5")
                        # Check the count of direct child div elements
                        try:
                            jaaab=jaaab.find_element(By.CSS_SELECTOR,'.optional-action-target-wrapper')
                            pjob=item.find_element(By.CSS_SELECTOR,'span[aria-hidden="true"]')
                            print(pjob.text," :Line 177 Companies\n")
                            if pjob.text in target_companies:
                                #past position
                               pjobea=item.find_element(By.CSS_SELECTOR,'div[class="pvs-list__outer-container"]')
                               pjobea=pjobea.find_element(By.CSS_SELECTOR,'span[aria-hidden="true"]')
                               past_job=pjobea.text 
                               print("Past Position",past_job,"\n")
                               dt=item.find_element(By.CSS_SELECTOR,'div[class="pvs-list__outer-container"]')
                               dt=dt.find_element(By.CSS_SELECTOR,'span[class="t-14 t-normal t-black--light"]')
                               past_date=dt.text
                               print("Past Date: ",past_date,"\n")
                               past_company=pjob.text
                               print("Past Company: ",past_company,"\n")
                        except TimeoutException:
                            job3=item.find_element(By.CSS_SELECTOR,'span[class="t-14 t-normal"]')
                            job3=job3.find_element(By.CSS_SELECTOR,'span[aria-hidden="true"]')
                            if job3.text in target_companies:
                               job4=item.find_element(By.CSS_SELECTOR,'span[aria-hidden="true"]')
                               past_job=job4.text
                               print("Past Job: ",past_job,"\n")
                               dt=item.find_element(By.CSS_SELECTOR,'span[class="t-14 t-normal t-black--light"]')
                               dt=dt.find_element(By.CSS_SELECTOR,'span[aria-hidden="true"]')
                               past_date=dt.text
                               print("Past Date: ",past_date,"\n")
                               past_company=job3.text
                               print("Past Company: ",past_company,"\n")
        except TimeoutException:
           past_job="None"
           print("Past Job: ",past_job,"\n")
           past_date="None"
           print("Past Date: ",past_date,"\n")
           past_company="None"
           print("Past Company: ",past_company,"\n") 
        
    # Close the tab
    driver.close()

    # Switch back to the original tab
    driver.switch_to.window(driver.window_handles[0])                                    

    # Return the extracted details
    return {
        "name": name,
        "current_job": current_job,
        "current_company": current_company,
        "about_data": about_data,
        "past_company": past_company,
        "past_job": past_job,
        "past_date":past_date,
        "contact_info": contact_info  
    }
  
driver=webdriver.Chrome()
driver.maximize_window()
# Open the LinkedIn login page
driver.get("https://www.linkedin.com/login")
driver.implicitly_wait(5)
login()
input("Say Something")
#Search
search=driver.find_element(By.CSS_SELECTOR,'button[class="search-global-typeahead__collapsed-search-button"]')
search.click()
search_typebox=driver.find_element(By.CSS_SELECTOR,'input[class="search-global-typeahead__input"]')
search_typebox.send_keys("sears")
search_typebox.send_keys(Keys.RETURN)
driver.implicitly_wait(15)

#Clicking people
people=driver.find_element(By.CSS_SELECTOR,'.artdeco-pill.artdeco-pill--slate.artdeco-pill--choice.artdeco-pill--2.search-reusables__filter-pill-button')
people.click()

#Clicking Filters
filters=driver.find_element(By.CSS_SELECTOR,'.artdeco-pill.artdeco-pill--slate.artdeco-pill--choice.artdeco-pill--2.search-reusables__filter-pill-button.search-reusables__all-filters-pill-button')
filters.click()
driver.implicitly_wait(5)
filters=driver.find_element(By.CSS_SELECTOR,'div[class="artdeco-modal-overlay artdeco-modal-overlay--layer-default artdeco-modal-overlay--is-top-layer search-reusables__side-panel-overlay ember-view"]')

arr_ay=["Sears Hometown and Outlet Stores, Inc.","Sears Home Improvements","Sears Home & Business Franchises","Sears Home Services","Shop Your Way (SYW)","Transformco","Sears","Sears Stores","Sears Canada","Sears Holdings Corporation"]
#adding company
filters=filters.find_element(By.CSS_SELECTOR,'.artdeco-modal__content.ember-view.display-flex.relative.mb4.flex-1.justify-center')
filters=filters.find_elements(By.CSS_SELECTOR,'li[class="search-reusables__secondary-filters-filter"]')
for past in filters:
    pd=past.find_element(By.CSS_SELECTOR,'h3[class="t-16 t-black t-bold inline-block"]')
   
    #to selct country
    if pd.text == "Locations":
        coun=past.find_element(By.CSS_SELECTOR,'ul[class="list-style-none relative display-flex flex-wrap list-style-none"]')
        coun=coun.find_elements(By.TAG_NAME,'li')
        for country in coun:
            count=country.find_element(By.CSS_SELECTOR,'label[class="search-reusables__value-label"]')
            count=count.find_element(By.CSS_SELECTOR,'span[class="t-14 t-black--light t-normal"]')
            if count.text == "United States":
                count.click()
                break
    
    #to select past companies    
    if pd.text == "Past company":
        driver.execute_script("arguments[0].scrollIntoView({ behavior: 'smooth', block: 'center' });", pd)
        for i in range(0,1):
            add_comnay=past.find_element(By.CSS_SELECTOR,'button[class="artdeco-button artdeco-button--muted artdeco-button--2 artdeco-button--tertiary ember-view reusable-search-filters-advanced-filters__add-filter-button"]')
            add_comnay.click()
            time.sleep(1)
        
            #Select companies
            com=past.find_element(By.CSS_SELECTOR,'input[placeholder="Add a company"]')
            com.send_keys(arr_ay[i])
            time.sleep(1)
            com.send_keys(Keys.CONTROL + Keys.ARROW_DOWN)
            com.send_keys(Keys.ENTER)
        dne=driver.find_element(By.CSS_SELECTOR,'button[class="reusable-search-filters-buttons search-reusables__secondary-filters-show-results-button artdeco-button artdeco-button--2 artdeco-button--primary ember-view"]')    
        dne.click()
        break
row=1
y=1
driver.implicitly_wait(5)
input("Password")
for x in range(1,100):
    # Now to filter out the person
    person=driver.find_element(By.CSS_SELECTOR,'.pv0.ph0.mb2.artdeco-card')
    person=person.find_element(By.CSS_SELECTOR,'ul[class="reusable-search__entity-result-list list-style-none"]')
    persons=person.find_elements(By.TAG_NAME,'li')
    leng=len(persons)
    
    for p in persons:
        nm=p.find_element(By.CSS_SELECTOR,'div[class="entity-result__content entity-result__divider pt3 pb3 t-12 t-black--light"]')
        nm=nm.find_element(By.CSS_SELECTOR,'.entity-result__title-text.t-16')
        if nm.text == "LinkedIn Member":
            continue
        else:
            nm=nm.find_element(By.CSS_SELECTOR,'a[class="app-aware-link "]')
            linkedin_url=nm.get_attribute("href")
            result = extract_linkedin_details(linkedin_url, arr_ay)
            sh1.cell(row, column=1, value=result["name"])
            sh1.cell(row, column=2, value=result["current_job"])
            sh1.cell(row, column=3, value=result["current_company"])
            sh1.cell(row, column=4, value=result["contact_info"])
            sh1.cell(row, column=5, value=result["about_data"])
            sh1.cell(row, column=6, value=result["past_job"])
            sh1.cell(row, column=7, value=result["past_date"])
            sh1.cell(row, column=8, value=result["past_company"])
            wb.save('C:\\Users\\User\\Desktop\\people1.xlsx')
            row+=1
            
    btn=driver.find_element(By.CSS_SELECTOR,'button[aria-label="Next"]')
    btn.click()
    y+=1
    print("Page Number",y)

driver.quit()        
        



